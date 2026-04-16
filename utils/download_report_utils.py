import datetime
import shutil
import time
from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from constants.constants import default_download_path
from constants.file_paths import template_file_path, output_directory
from helper.chrome_helper import ChromeHelper
from helper.crawler_helper import crawler_helper
from helper.logging_helper import logger
from model.eams_wo import EAMSWorkOrder


def generate_timestamped_filename(prefix: str = "result",
                                  extension: str = ".xls") -> str:
    """
    Generates a dynamically named file string with a current timestamp.

    By using parameters with default values, this function adheres more
    closely to the Open/Closed Principle, allowing it to be reused for
    different file types and prefixes without modifying its internal logic.

    Args:
        prefix (str): The prefix of the file name. Defaults to "result".
        extension (str): The file extension. Defaults to ".xls".

    Returns:
        str: A formatted filename string (e.g., "result_2026_04_16_095400.xls").
    """
    # Defensive programming: ensure the extension always starts with a dot
    if not extension.startswith('.'):
        extension = f".{extension}"

    current_time = datetime.datetime.now().strftime("%Y_%m_%d_%H%M%S")
    output_file_name = f"{prefix}_{current_time}{extension}"

    return output_file_name


def rename_file(downloads_path: Path,
                new_file_path: str):
    files = list(downloads_path.glob("*.xls"))
    if not files:
        logger.warning("No Excel files found in the Downloads folder.")
    else:
        latest_file = max(files, key=lambda p: p.stat().st_mtime)
        try:
            latest_file.rename(new_file_path)
            logger.info(f"Renamed: {latest_file.name}")
            logger.info(f"To:      {new_file_path}")
        except FileExistsError:
            logger.error(f"Error: A file named {new_file_path} already exists.")


def read_eams_record(file_path) -> List[EAMSWorkOrder]:
    # 1. Use read_html because the file is actually an HTML table
    # This returns a list of dataframes; usually, the data is in the first one [0]
    try:
        dfs = pd.read_html(file_path, header=0)
        df = dfs[0]
    except Exception as e:
        logger.info(f"Error parsing HTML table: {e}")
        return []

    work_orders = []

    # 2. Iterate through the rows of the dataframe
    for _, row in df.iterrows():
        # Helper to safely convert strings to datetime
        def to_dt(val):
            if pd.isna(val) or str(val).strip() == "":
                return None
            try:
                return pd.to_datetime(val).to_pydatetime()
            except:
                return None

        # 3. Map the dataframe columns to your Dataclass fields
        # Note: df.iloc[n] uses the column index (0, 1, 2...)
        wo = EAMSWorkOrder(
            work_order_id=int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
            description=str(row.iloc[1]),
            work_group=str(row.iloc[2]),
            work_type=str(row.iloc[3]),
            asset=str(row.iloc[4]),
            asset_classification=str(row.iloc[5]),
            location=str(row.iloc[6]),
            status=str(row.iloc[7]),
            job_plan=str(row.iloc[8]),
            target_start_datetime=to_dt(row.iloc[9]),
            target_finish_datetime=to_dt(row.iloc[10]),
            actual_start_datetime=to_dt(row.iloc[13]),  # Optional
            actual_finish_datetime=to_dt(row.iloc[14])  # Optional
        )
        work_orders.append(wo)
    return work_orders


def write_to_template(output_file_path: str,
                      wo_records: List[EAMSWorkOrder]):
    # output_file_path = os.path.join(output_directory, output_file_name)

    try:
        wb = openpyxl.load_workbook(template_file_path)
        if "Sheet9" in wb.sheetnames:
            sheet = wb["Sheet9"]
        else:
            logger.info("Sheet9 not found. Creating it now...")
            sheet = wb.create_sheet("Sheet9")
        for row in wo_records:
            sheet.append(row.to_list_data())
            wb.save(output_file_path)
            logger.info(f"Success! A sample row has been added to {output_file_path} on 'Sheet9'.")

    except FileNotFoundError:
        logger.error(f"Error: The file '{output_file_path}' was not found. Check the file name and path.")
    except PermissionError:
        logger.error(f"Error: Could not save the file. Please close '{output_file_path}' in Excel and try again.")


def wait_and_rename_latest(downloads_path: Path,
                           new_file_path: str,
                           timeout=30) -> bool:
    """Waits up to 'timeout' seconds for an .xls file to appear, then renames it."""
    start_time = time.time()

    while time.time() - start_time < timeout:
        files = list(downloads_path.glob("*.xls"))
        if files:
            latest_file = max(files, key=lambda p: p.stat().st_mtime)
            # Ensure it's not a temp download file
            if not latest_file.name.endswith('.crdownload'):
                try:
                    latest_file.rename(new_file_path)
                    return True
                except FileExistsError:
                    return True  # Or handle as needed
        time.sleep(2)  # Polling interval
    return False


def move_to_archive(file_to_move: Path,
                    destination_folder: Path):
    """Moves the processed file to an archive directory."""
    try:
        if not destination_folder.exists():
            destination_folder.mkdir(parents=True, exist_ok=True)

        target_path = destination_folder / file_to_move.name
        shutil.move(str(file_to_move), str(target_path))
        logger.info(f"Housekeeping: File moved to {target_path}")
    except Exception as e:
        logger.warning(f"Housekeeping failed: {e}")


def clear_downloads(downloads_path: Path):
    """
    Deletes all files in the specified directory.
    Does not delete sub-folders to avoid accidental data loss.
    """
    logger.info(f"Cleaning up Downloads folder: {downloads_path}")

    if not downloads_path.exists():
        logger.warning("Downloads path does not exist. Skipping cleanup.")
        return

    file_count = 0
    for item in downloads_path.iterdir():
        # Only delete files, ignore sub-directories
        if item.is_file():
            try:
                item.unlink()
                file_count += 1
            except PermissionError:
                logger.warning(f"Could not delete {item.name} - File is currently in use.")
            except Exception as e:
                logger.error(f"Error deleting {item.name}: {e}")

    logger.info(f"Cleanup complete. Removed {file_count} files.")


def download_report() -> None:
    logger.info("initialization")
    output_file_name = generate_timestamped_filename()

    downloads_dir = Path.home() / default_download_path
    output_dir = Path(output_directory)
    housekeeping_dir = output_dir / "Archive"
    download_file_path = downloads_dir / output_file_name
    output_file_path = output_dir / output_file_name

    clear_downloads(downloads_dir)
    try:
        logger.info("Downloading today's CM/PM Report for LAR...")
        (
            crawler_helper
            .set_chrome_helper(ChromeHelper())
            .login()
            .go_to_wo_tracking_page()
            .search_and_download_reports()
        )

        success = wait_and_rename_latest(downloads_dir, str(download_file_path))
        if not success:
            raise FileNotFoundError("The report was not found in the Downloads folder after the crawler finished.")
        logger.info("Reading record from downloaded HTML-XLS...")

        records = read_eams_record(str(download_file_path))
        if not records:
            logger.warning("File was read, but no work order records were found.")
        else:
            write_to_template(str(output_file_path), records)
        move_to_archive(download_file_path, housekeeping_dir)

    except Exception as e:
        # Graceful User Notification
        logger.error("-" * 50)
        logger.error(f"CRITICAL ERROR IN FLOW: {str(e)}")
        logger.error("Please check your internet connection or EAMS credentials.")
        logger.error("-" * 50)
        # You could also trigger a system popup or email here
    finally:
        logger.info("CMCR Process Finished.")
