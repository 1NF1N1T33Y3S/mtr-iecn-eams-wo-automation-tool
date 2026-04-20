import shutil
import time
import sys
from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from constants.constants import PROJECT_DOWNLOAD_DIR
from constants.email_configs import LAR_RECIPIENTS, EMAIL_SUBJECT, TEST_RECIPIENTS
from constants.file_paths import template_file_path
from exceptions.exceptions import EAMSReportNotFoundError
from helper.chrome_helper import ChromeHelper
from helper.crawler_helper import crawler_helper
from helper.email_manager import email_handler
from helper.logging_helper import logger
from model.eams_wo import EAMSWorkOrder
from model.email import get_no_work_order_email, Email
from utils.utils import generate_timestamped_filename


def wait_and_rename_latest(downloads_path: Path,
                           target_path: Path,
                           timeout: int = 30) -> None:
    """
    Waits up to 'timeout' seconds for an .xls file to appear, then renames it.
    Raises a FileNotFoundError if the file does not appear within the timeout.
    """
    start_time = time.time()

    while time.time() - start_time < timeout:
        files = list(downloads_path.glob("*.xls"))
        if files:
            latest_file = max(files, key=lambda p: p.stat().st_mtime)

            # Ensure it's not a temp download file
            if not latest_file.name.endswith('.crdownload'):
                try:
                    # Overwrite if target already exists (avoids FileExistsError on Windows)
                    if target_path.exists():
                        target_path.unlink()

                    latest_file.rename(target_path)
                    logger.info(f"Successfully found and renamed {latest_file.name}")
                    return  # Exit function on success
                except Exception as e:
                    logger.warning(f"Attempted to rename but encountered an issue (retrying): {e}")

        time.sleep(2)  # Polling interval

    # If the loop finishes without returning, the timeout was reached
    raise EAMSReportNotFoundError(str(downloads_path))


def read_eams_record(file_path: Path) -> List[EAMSWorkOrder]:
    """Reads HTML-based Excel files and parses them into EAMSWorkOrder objects."""
    try:
        dfs = pd.read_html(str(file_path), header=0)
        df = dfs[0]
    except Exception as e:
        logger.error(f"Error parsing HTML table: {e}")
        return []

    work_orders = []

    def to_dt(val):
        if pd.isna(val) or str(val).strip() == "":
            return None
        try:
            return pd.to_datetime(val).to_pydatetime()
        except Exception:
            return None

    for _, row in df.iterrows():
        wo = EAMSWorkOrder(
            work_order_id=int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
            description=str(row.iloc[1]),
            work_group=str(row.iloc[2]),
            work_type=str(row.iloc[3]),
            asset=str(row.iloc[4]),
            asset_classification=str(row.iloc[5]),
            location=str(row.iloc[6]),
            status=str(row.iloc[7]),
            job_plan=str(row.iloc[8]),
            target_start_datetime=to_dt(row.iloc[9]),
            target_finish_datetime=to_dt(row.iloc[10]),
            actual_start_datetime=to_dt(row.iloc[13]),
            actual_finish_datetime=to_dt(row.iloc[14])
        )
        work_orders.append(wo)

    return work_orders


def export_records_to_template(output_file_path: Path,
                               records: List[EAMSWorkOrder]) -> None:
    """Exports a list of EAMSWorkOrder objects to the specified Excel template."""
    if not records:
        logger.warning("No work order records were found. Skipping export.")
        return

    try:
        wb = openpyxl.load_workbook(template_file_path)
        sheet_name = "Sheet9"

        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            logger.info(f"'{sheet_name}' not found. Creating it now...")
            sheet = wb.create_sheet(sheet_name)

        for row in records:
            sheet.append(row.to_list_data())

        wb.save(output_file_path)
        logger.info(f"Successfully wrote {len(records)} records to '{output_file_path}'.")

    except FileNotFoundError:
        logger.error(f"Error: The template file '{template_file_path}' was not found.")
        raise
    except PermissionError:
        logger.error(f"Error: Could not save to '{output_file_path}'. Please ensure it is closed.")
        raise


def move_to_archive(file_to_move: Path,
                    destination_folder: Path) -> None:
    """Moves the processed file to an archive directory."""
    try:
        destination_folder.mkdir(parents=True, exist_ok=True)
        target_path = destination_folder / file_to_move.name
        shutil.move(str(file_to_move), str(target_path))
        logger.info(f"Housekeeping: File moved to {target_path}")
    except Exception as e:
        logger.warning(f"Housekeeping failed: {e}")


def clear_folder(folder_path: Path) -> None:
    """Deletes all files in the specified directory, leaving sub-folders intact."""
    logger.info(f"Cleaning up Downloads folder: {folder_path}")

    if not folder_path.exists():
        logger.warning("Downloads path does not exist. Skipping cleanup.")
        return

    file_count = 0
    for item in folder_path.iterdir():
        if item.is_file():
            try:
                item.unlink()
                file_count += 1
            except PermissionError:
                logger.warning(f"Could not delete {item.name} - File is currently in use.")
            except Exception as e:
                logger.error(f"Error deleting {item.name}: {e}")

    logger.info(f"Cleanup complete. Removed {file_count} files.")


def process_downloaded_report(
        downloads_dir: Path,
        download_file_path: Path,
        output_file_path: Path,
        housekeeping_dir: Path
) -> List[EAMSWorkOrder]:
    """Main workflow for processing the downloaded report."""
    wait_and_rename_latest(downloads_dir, download_file_path)

    logger.info("Reading records from downloaded file...")
    records = read_eams_record(download_file_path)
    export_records_to_template(output_file_path, records)
    move_to_archive(output_file_path, housekeeping_dir)
    logger.info("process_downloaded_report completed")
    return records


def download_report() -> None:
    logger.info("downloading report")
    logger.info("Downloading today's CM/PM Report for LAR...")
    (
        crawler_helper
        .set_chrome_helper(ChromeHelper())
        .login()
        .go_to_wo_tracking_page()
        .search_and_download_reports()
    )


def download_report_pipeline(
        output_file_name: str) -> List[EAMSWorkOrder]:
    logger.info("Starting the report generation pipeline...")
    downloads_dir = Path(PROJECT_DOWNLOAD_DIR)
    download_file_path = downloads_dir / output_file_name
    output_dir = Path("output")
    output_file_path = output_dir / output_file_name
    housekeeping_dir = Path("archive/")

    try:
        logger.info("Initiating report download...")
        clear_folder(downloads_dir)
        clear_folder(output_dir)
        download_report()

        logger.info("Processing the downloaded report...")
        return process_downloaded_report(
            downloads_dir=downloads_dir,
            download_file_path=download_file_path,
            output_file_path=output_file_path,
            housekeeping_dir=housekeeping_dir
        )

    except EAMSReportNotFoundError as e:
        logger.error(str(e))
        fail_body = get_no_work_order_email()
        fail_email = Email(LAR_RECIPIENTS, EMAIL_SUBJECT, fail_body)
        email_handler.send_email(fail_email)
    except Exception as e:
        logger.error(f"A critical error occurred during the report downloading pipeline execution: {e}", exc_info=True)
    finally:
        logger.error("Ending the application")
        sys.exit(1)
