import openpyxl

from typing import Self

from openpyxl.worksheet.worksheet import Worksheet


class ExcelHelper:
    def __init__(self):
        self.file_path: str = ""
        self.sheet_name: str = ""
        self.work_book: openpyxl.Workbook = None
        self.work_sheet: Worksheet = None
        self.max_rows: int = 0

    def set_file_path(self, file_path: str) -> Self:
        self.file_path: str = file_path
        return self

    def set_sheet_name(self, sheet_name: str) -> Self:
        self.sheet_name: str = sheet_name
        return self

    def set_max_row(self, rows: int) -> Self:
        self.max_rows: int = rows
        return self

    def read_excel(self) -> Self:
        self.work_book = openpyxl.load_workbook(self.file_path)
        self.work_sheet = self.work_book[self.sheet_name]
        return self

    def get_row_by_column(self, column_index: int, value: str) -> int:
        row_count = 3
        for row in self.work_sheet.iter_rows(min_row=row_count, max_row=self.max_rows + 1):
            # print(f"{row[column_index].value=}")
            column_value = str(row[column_index].value)
            if column_value == value:
                return row_count
            row_count += 1
        return -1

    def write(self, column_name: str, row_index: int, value: str):
        print(f"{column_name}, {row_index}: {value}")
        first_row = self.work_sheet[1]
        header = [cell.value for cell in first_row]
        column_index = header.index(column_name) + 1
        self.work_sheet.cell(row=row_index, column=column_index).value = value
        # self.work_book.save(self.file_path)
        print("write done")

    def save(self):
        self.work_book.save(self.file_path)
        print("save done")
