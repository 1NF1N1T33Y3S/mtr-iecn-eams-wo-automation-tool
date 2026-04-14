import os
import openpyxl

# 1. Specify your file name
folder_path = r"C:\Users\leecamkf\PycharmProjects\iecn-eams-workorder-automation-tool\input"
file_name = 'test_1.xlsx'
input_file_name = "cmwo-template.xlsx"

input_file_path = os.path.join(folder_path, input_file_name)
output_file_path = os.path.join(folder_path, file_name)

try:
    wb = openpyxl.load_workbook(input_file_path)
    if "Sheet9" in wb.sheetnames:
        sheet = wb["Sheet9"]
    else:
        print("Sheet9 not found. Creating it now...")
        sheet = wb.create_sheet("Sheet9")
    sample_row = ["AELnTCL",
                  "",
                  "",
                  "",
                  ""]
    sheet.append(sample_row)
    wb.save(output_file_path)
    print(f"Success! A sample row has been added to {output_file_path} on 'Sheet9'.")

except FileNotFoundError:
    print(f"Error: The file '{output_file_path}' was not found. Check the file name and path.")
except PermissionError:
    print(f"Error: Could not save the file. Please close '{output_file_path}' in Excel and try again.")
