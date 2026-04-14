import os, time
from pathlib import Path
from datetime import datetime
from typing import List

import openpyxl
import pandas as pd

from constants.constants import default_download_path
from constants.file_paths import iecc_centralized_log_file_path, template_file_path, output_directory
from helper.chrome_helper import ChromeHelper
from helper.crawler_helper import crawler_helper
from helper.excel_helper import ExcelHelper, excel_helper
from helper.logging_helper import logger
from model.eams_wo import EAMSWorkOrder
from model.work_order import WorkOrder


def output_result_to_excel(excel: ExcelHelper, wos: List[WorkOrder]):
    for w in wos:
        logger.info(f"{w.id=} {w.wo_id=} {w.actual_start_date=} {w.actual_finish_date=}")
        r = excel.get_row_by_column(0, w.id)
        excel.write("EAMS WO Completed", r, w.job_status)
        excel.write("EAMS WO Failure Message", r, w.execution_error_message)


def rename_file(downloads_path: Path,
                new_file_path: str):
    files = list(downloads_path.glob("*.xls"))
    if not files:
        logger.warning("No Excel files found in the Downloads folder.")
    else:
        latest_file = max(files, key=lambda p: p.stat().st_mtime)
        try:
            latest_file.rename(new_file_path)
            logger.info(f"Renamed: {latest_file.name}")
            logger.info(f"To:      {new_file_path}")
        except FileExistsError:
            logger.error(f"Error: A file named {new_file_path} already exists.")


def read_eams_record(file_path):
    # 1. Use read_html because the file is actually an HTML table
    # This returns a list of dataframes; usually, the data is in the first one [0]
    try:
        dfs = pd.read_html(file_path, header=0)
        df = dfs[0]
    except Exception as e:
        logger.info(f"Error parsing HTML table: {e}")
        return []

    work_orders = []

    # 2. Iterate through the rows of the dataframe
    for _, row in df.iterrows():

        # Helper to safely convert strings to datetime
        def to_dt(val):
            if pd.isna(val) or str(val).strip() == "":
                return None
            try:
                return pd.to_datetime(val).to_pydatetime()
            except:
                return None

        # 3. Map the dataframe columns to your Dataclass fields
        # Note: df.iloc[n] uses the column index (0, 1, 2...)
        wo = EAMSWorkOrder(
            work_order_id=int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
            description=str(row.iloc[1]),
            work_group=str(row.iloc[2]),
            work_type=str(row.iloc[3]),
            asset=str(row.iloc[4]),
            asset_classification=str(row.iloc[5]),
            location=str(row.iloc[6]),
            status=str(row.iloc[7]),
            job_plan=str(row.iloc[8]),
            target_start_datetime=to_dt(row.iloc[9]),
            target_finish_datetime=to_dt(row.iloc[10]),
            actual_start_datetime=to_dt(row.iloc[13]),  # Optional
            actual_finish_datetime=to_dt(row.iloc[14])  # Optional
        )
        work_orders.append(wo)
    return work_orders


def write_to_template(output_file_name: str,
                      wo_records: List[EAMSWorkOrder]):
    output_file_path = os.path.join(output_directory, output_file_name)

    try:
        wb = openpyxl.load_workbook(template_file_path)
        if "Sheet9" in wb.sheetnames:
            sheet = wb["Sheet9"]
        else:
            logger.info("Sheet9 not found. Creating it now...")
            sheet = wb.create_sheet("Sheet9")
        for row in wo_records:
            sheet.append(row.to_list_data())
            wb.save(output_file_path)
            logger.info(f"Success! A sample row has been added to {output_file_path} on 'Sheet9'.")

    except FileNotFoundError:
        logger.error(f"Error: The file '{output_file_path}' was not found. Check the file name and path.")
    except PermissionError:
        logger.error(f"Error: Could not save the file. Please close '{output_file_path}' in Excel and try again.")


def main():
    logger.info('CMCR Automation Tool')
    logger.info("Downloading today's CM/PM Report for LAR")
    (
        crawler_helper
        .set_chrome_helper(ChromeHelper())
        .login()
        .go_to_wo_tracking_page()
        .search_and_download_reports()
    )

    logger.info("Renaming the downloaded report")
    current_time = datetime.now().strftime("%Y_%m_%d_%H%M%S")
    new_filename = f"{current_time}.xls"
    downloads_path = Path.home() / default_download_path
    eams_wo_order_path = downloads_path / new_filename

    rename_file(downloads_path, str(eams_wo_order_path))
    records = read_eams_record(str(eams_wo_order_path))
    write_to_template("output.xlsx", records)


if __name__ == '__main__':
    main()
